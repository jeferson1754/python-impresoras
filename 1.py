import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Obtener la fecha y hora actual
timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

num_version = "139.0.7258.139"

CHROMEDRIVER_PATH = ChromeDriverManager(
    driver_version=num_version).install()

# Configura el driver y el navegador


def configurar_driver(num_version: str):
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')  # Ãºtil en algunos entornos Linux
    # mejora en entornos con pocos recursos
    options.add_argument('--disable-dev-shm-usage')
    # asegura resoluciÃ³n adecuada en modo headless
    # options.add_argument('--window-size=1920,1080')
    # options.add_argument('--window-size=1080,1080')
    options.add_argument('--ignore-certificate-errors')  # Ignorar errores SSL
    options.add_argument('--ignore-ssl-errors')

    driver = webdriver.Chrome(
        service=Service(CHROMEDRIVER_PATH),
        options=options
    )

    return driver

# FunciÃ³n para formatear la IP


def format_ip(ip):
    if pd.isna(ip) or not ip.strip():
        return None  # Retorna None si la IP estÃ¡ vacÃ­a o solo tiene espacios

    ip = re.sub(r'\D', '', ip)

    if len(ip) == 12:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 11:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 10:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 9:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 8:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    elif len(ip) == 7:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    else:
        return ip

# FunciÃ³n para configurar el ancho de las columnas


def get_column_widths(ws):
    widths = {}
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        widths[column] = max_length + 2  # Adding a bit more space
    return widths

# Funciones para preservar y aplicar fÃ³rmulas y formatos


def preserve_formulas_and_formats(input_file):
    wb = load_workbook(input_file, data_only=False)
    formulas = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = {cell.coordinate: cell.formula if cell.data_type ==
                          'f' else cell.value for row in ws.iter_rows() for cell in row}
        formulas[sheet_name] = sheet_formulas
    return formulas

# Aplicar fÃ³rmulas y formatos preservados


def apply_formulas_and_formats(output_file, formulas):
    wb = load_workbook(output_file, data_only=False)
    for sheet_name, sheet_formulas in formulas.items():
        ws = wb[sheet_name]
        for cell_coord, formula in sheet_formulas.items():
            cell = ws[cell_coord]
            if formula is not None:
                if isinstance(formula, str) and formula.startswith('='):
                    cell.formula = formula
                else:
                    cell.value = formula
    wb.save(output_file)


def procesar_impresoras_hp(file_path, output_file):

    def clean_percentage(value: str) -> str:
        try:
            if isinstance(value, str):
                value = value.replace('%', '').strip()
            return f"{int(round(float(value)))}%"
        except ValueError:
            return ""

    def fetch_data_from_url(ip, options):
        url = f"http://{ip}" if ip else None
        if not url:
            return {"IP": ip, "Toner Negro": "", "Kit Mant.": "", "Kit Alim.": "", 'Estado': '', 'Marca de Tiempo': ""}

        print(f"Procesando URL: {url}")
        driver = configurar_driver(num_version)

        try:

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "SupplyName0"))
            )

            # Toner negro
            toner_negro = driver.find_element(By.ID, "SupplyGauge0").text
            # print(f"TÃ³ner Negro: {toner_negro}")

            # Kit de mantenimiento
            kit_mantenimiento = driver.find_element(By.ID, "SupplyGauge1").text
            # print(f"Kit de mantenimiento: {kit_mantenimiento}")

            # Kit alimentador documentos
            kit_alimentador = driver.find_element(By.ID, "SupplyGauge2").text
            # print(f"Kit alimentador documentos: {kit_alimentador}")

            return {
                "IP": ip,
                "Toner Negro": toner_negro,
                "Kit Mant.": kit_mantenimiento,
                "Kit Alim.": kit_alimentador,
                'Estado': 'OK' if toner_negro or kit_mantenimiento or kit_alimentador else 'No disponible',
                'Marca de Tiempo': timestamp
            }
        except (NoSuchElementException):
            return {"IP": ip, "Toner Negro": "", "Kit Mant.": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except TimeoutException:
            return {"IP": ip, "Toner Negro": "", "Kit Mant.": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except WebDriverException:
            print(f"Timeout al intentar conectar con {url}")
            return {"IP": ip, "Toner Negro": "", "Kit Mant.": "", "Kit Alim.": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
        finally:
            driver.quit()

    # Leer las hojas del archivo Excel
    sheets = pd.read_excel(file_path, sheet_name=None)
    wb = load_workbook(file_path)
    column_widths = {sheet_name: get_column_widths(
        wb[sheet_name]) for sheet_name in wb.sheetnames}

    # Procesar la hoja 'HP Admin'
    df_original = sheets['HP Admin']
    df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(
            fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    # Verificar que las columnas 'IP' existan antes de hacer el merge
    if 'IP' not in df_original.columns or 'IP' not in df_results.columns:
        raise KeyError("'IP' column is missing in one of the DataFrames.")

    # Fusionar los resultados
    df_updated = df_original.merge(
        df_results, on='IP', how='left', suffixes=('', '_new')
    )

    # Restablecer columnas NaN
    columns = ['Toner Negro', 'Kit Mant.',
               'Kit Alim.', 'Estado', 'Marca de Tiempo']
    df_updated[columns] = df_updated[columns].fillna('')

    mask_ok = df_updated['Estado_new'] == 'OK'
    columns = ['Toner Negro', 'Kit Mant.', 'Kit Alim.']
    for col in columns:
        df_updated[col] = df_updated[col].astype(str)
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok,
                                                      f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(
        df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(
        df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in columns +
                       ['Estado', 'Marca de Tiempo']]
    df_updated.drop(columns=[
                    col for col in columns_to_drop if col in df_updated.columns], inplace=True)

    # Guardar el DataFrame actualizado
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(
            writer, sheet_name='HP Admin', index=False)
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'HP Admin':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar fÃ³rmulas y formatos preservados
    formulas = preserve_formulas_and_formats(file_path)
    apply_formulas_and_formats(output_file, formulas)
    # FUNCIONA
    registrar_historico(file_path, df_updated)


def procesar_impresoras_hp_grandes(file_path, output_file):

    def clean_percentage(value: str) -> str:
        try:
            if isinstance(value, str):
                value = value.replace('%', '').strip()
            return f"{int(round(float(value)))}%"
        except ValueError:
            return ""

    def fetch_data_from_url(ip, options):
        url = f"http://{ip}" if ip else None
        if not url:
            return {"IP": ip, "Toner Negro": "", 'Estado': '', 'Marca de Tiempo': ""}

        print(f"Procesando URL: {url}")
        driver = configurar_driver(num_version)

        try:

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "SupplyName0"))
            )

            # Toner negro
            toner_negro = driver.find_element(By.ID, "SupplyGauge0").text
            # print(f"TÃ³ner Negro: {toner_negro}")

            return {
                "IP": ip,
                "Toner Negro": toner_negro,
                'Estado': 'OK' if toner_negro else 'No disponible',
                'Marca de Tiempo': timestamp
            }
        except (NoSuchElementException):
            return {"IP": ip, "Toner Negro": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except TimeoutException:
            return {"IP": ip, "Toner Negro": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except WebDriverException:
            print(f"Timeout al intentar conectar con {url}")
            return {"IP": ip, "Toner Negro": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
        finally:
            driver.quit()

    # Leer las hojas del archivo Excel
    sheets = pd.read_excel(file_path, sheet_name=None)
    wb = load_workbook(file_path)
    column_widths = {sheet_name: get_column_widths(
        wb[sheet_name]) for sheet_name in wb.sheetnames}

    # Procesar la hoja 'HP Planta'
    df_original = sheets['HP Planta']
    df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(
            fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    # Verificar que las columnas 'IP' existan antes de hacer el merge
    if 'IP' not in df_original.columns or 'IP' not in df_results.columns:
        raise KeyError("'IP' column is missing in one of the DataFrames.")

    # Fusionar los resultados
    df_updated = df_original.merge(
        df_results, on='IP', how='left', suffixes=('', '_new')
    )

    # Restablecer columnas NaN
    columns = ['Toner Negro', 'Estado', 'Marca de Tiempo']
    df_updated[columns] = df_updated[columns].fillna('')

    mask_ok = df_updated['Estado_new'] == 'OK'
    columns = ['Toner Negro']
    for col in columns:
        df_updated[col] = df_updated[col].astype(str)
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok,
                                                      f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(
        df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(
        df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in columns +
                       ['Estado', 'Marca de Tiempo']]
    df_updated.drop(columns=[
                    col for col in columns_to_drop if col in df_updated.columns], inplace=True)

    # Guardar el DataFrame actualizado
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(
            writer, sheet_name='HP Planta', index=False)
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'HP Planta':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar fÃ³rmulas y formatos preservados
    formulas = preserve_formulas_and_formats(file_path)
    apply_formulas_and_formats(output_file, formulas)
    # REVISANDO
    registrar_historico(file_path, df_updated)


def procesar_color_admin(file_path, output_file):

    def clean_percentage(value: str) -> str:
        try:
            if isinstance(value, str):
                value = value.replace('%', '').strip()
            return f"{int(round(float(value)))}%"
        except (ValueError, TypeError):
            return "0%"

    def fetch_data_from_url(ip, options):
        url = f"http://{ip}" if ip else None
        if not url:
            return {"IP": ip, "Toner Negro": "", "Toner Cian": "", "Toner Magenta": "", "Toner Amarillo": "",  'Estado': '', 'Marca de Tiempo': ""}

        print(f"Procesando URL: {url}")
        driver = configurar_driver(num_version)

        try:

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "SupplyName0"))
            )

            # Toner negro
            toner_negro = driver.find_element(By.ID, "SupplyGauge0").text
            toner_cian = driver.find_element(By.ID, "SupplyGauge1").text
            toner_magenta = driver.find_element(By.ID, "SupplyGauge2").text
            toner_amarillo = driver.find_element(By.ID, "SupplyGauge3").text

            '''
            print(f"TÃ³ner Negro: {toner_negro}")
            print(f"TÃ³ner Cian: {toner_cian}")
            print(f"TÃ³ner Magenta: {toner_magenta}")
            print(f"TÃ³ner Amarillo: {toner_amarillo}")
            '''

            return {
                "IP": ip,
                "Toner Negro": toner_negro,
                "Toner Cian": toner_cian,
                "Toner Magenta": toner_magenta,
                "Toner Amarillo": toner_amarillo,
                'Estado': 'OK' if toner_negro else 'No disponible',
                'Marca de Tiempo': timestamp
            }
        except (NoSuchElementException):
            return {"IP": ip, "Toner Negro": "", "Toner Cian": "", "Toner Magenta": "", "Toner Amarillo": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except TimeoutException:
            return {"IP": ip, "Toner Negro": "", "Toner Cian": "", "Toner Magenta": "", "Toner Amarillo": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except WebDriverException:
            print(f"Timeout al intentar conectar con {url}")
            return {"IP": ip, "Toner Negro": "", "Toner Cian": "", "Toner Magenta": "", "Toner Amarillo": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
        finally:
            driver.quit()

    # Leer las hojas del archivo Excel
    sheets = pd.read_excel(file_path, sheet_name=None)
    wb = load_workbook(file_path)
    column_widths = {sheet_name: get_column_widths(
        wb[sheet_name]) for sheet_name in wb.sheetnames}

    # Procesar la hoja 'Color Admin'
    df_original = sheets['Color Admin']
    df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(
            fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    # Verificar que las columnas 'IP' existan antes de hacer el merge
    if 'IP' not in df_original.columns or 'IP' not in df_results.columns:
        raise KeyError("'IP' column is missing in one of the DataFrames.")

    # Fusionar los resultados
    df_updated = df_original.merge(
        df_results, on='IP', how='left', suffixes=('', '_new')
    )

    # Restablecer columnas NaN
    columns = ['Toner Negro', 'Toner Cian', 'Toner Magenta',
               'Toner Amarillo', 'Estado', 'Marca de Tiempo']
    df_updated[columns] = df_updated[columns].fillna('')

    mask_ok = df_updated['Estado_new'] == 'OK'
    columns = ['Toner Negro', 'Toner Cian', 'Toner Magenta', 'Toner Amarillo']

    # ðŸ”¹ Crear las columnas *_new si no existen
    for col in columns:
        new_col = f"{col}_new"
        if new_col not in df_updated.columns:
            df_updated[new_col] = None

    # ðŸ”¹ Actualizar valores
    for col in columns:
        df_updated[col] = df_updated[col].astype(str)
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok,
                                                      f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(
        df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(
        df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in columns +
                       ['Estado', 'Marca de Tiempo']]
    df_updated.drop(columns=[
                    col for col in columns_to_drop if col in df_updated.columns], inplace=True)

    # Guardar el DataFrame actualizado
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(
            writer, sheet_name='Color Admin', index=False)
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'Color Admin':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar fÃ³rmulas y formatos preservados
    formulas = preserve_formulas_and_formats(file_path)
    apply_formulas_and_formats(output_file, formulas)
    registrar_historico(output_file, df_updated)


def procesar_planta(file_path, output_file):

    def clean_percentage(value: str) -> str:
        try:
            if isinstance(value, str):
                value = value.replace('%', '').strip()
            return f"{int(round(float(value)))}%"
        except ValueError:
            return ""

    def fetch_data_from_url(ip, options):
        url = f"http://{ip}" if ip else None
        if not url:
            return {"IP": ip, "Toner Negro": "", "Kit Alim.": "", 'Estado': '', 'Marca de Tiempo': ""}

        print(f"Procesando URL: {url}")
        driver = configurar_driver(num_version)

        try:

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "SupplyName0"))
            )

            # Toner negro
            toner_negro = driver.find_element(By.ID, "SupplyGauge0").text
            # print(f"TÃ³ner Negro: {toner_negro}")

            # Kit de mantenimiento
            kit_alimentador = driver.find_element(By.ID, "SupplyGauge1").text
            # print(f"Kit de mantenimiento: {kit_mantenimiento}")

            return {
                "IP": ip,
                "Toner Negro": toner_negro,
                "Kit Alim.": kit_alimentador,
                'Estado': 'OK' if toner_negro or kit_alimentador else 'No disponible',
                'Marca de Tiempo': timestamp
            }
        except (NoSuchElementException):
            return {"IP": ip, "Toner Negro": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except TimeoutException:
            return {"IP": ip, "Toner Negro": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except WebDriverException:
            print(f"Timeout al intentar conectar con {url}")
            return {"IP": ip, "Toner Negro": "", "Kit Alim.": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
        finally:
            driver.quit()

    # Leer las hojas del archivo Excel
    sheets = pd.read_excel(file_path, sheet_name=None)
    wb = load_workbook(file_path)
    column_widths = {sheet_name: get_column_widths(
        wb[sheet_name]) for sheet_name in wb.sheetnames}

    # Procesar la hoja 'HP Planta - 2'
    df_original = sheets['HP Planta - 2']
    df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(
            fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    # Verificar que las columnas 'IP' existan antes de hacer el merge
    if 'IP' not in df_original.columns or 'IP' not in df_results.columns:
        raise KeyError("'IP' column is missing in one of the DataFrames.")

    # Fusionar los resultados
    df_updated = df_original.merge(
        df_results, on='IP', how='left', suffixes=('', '_new')
    )

    # Restablecer columnas NaN
    columns = ['Toner Negro', 'Kit Alim.', 'Estado', 'Marca de Tiempo']
    df_updated[columns] = df_updated[columns].fillna('')

    mask_ok = df_updated['Estado_new'] == 'OK'
    columns = ['Toner Negro', 'Kit Alim.']
    for col in columns:
        df_updated[col] = df_updated[col].astype(str)
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok,
                                                      f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(
        df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(
        df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in columns +
                       ['Estado', 'Marca de Tiempo']]
    df_updated.drop(columns=[
                    col for col in columns_to_drop if col in df_updated.columns], inplace=True)

    # Guardar el DataFrame actualizado
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(
            writer, sheet_name='HP Planta - 2', index=False)
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'HP Planta - 2':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar fÃ³rmulas y formatos preservados
    formulas = preserve_formulas_and_formats(file_path)
    apply_formulas_and_formats(output_file, formulas)
    # NO FUNCIONA
    registrar_historico(output_file, df_updated)


def procesar_color_planta(file_path, output_file):

    def clean_percentage(value: str) -> str:
        try:
            if isinstance(value, str):
                value = value.replace('%', '').strip()
            return f"{int(round(float(value)))}%"
        except (ValueError, TypeError):
            return "0%"

    def fetch_data_from_url(ip, options):
        url = f"http://{ip}" if ip else None
        if not url:
            return {"IP": ip, "Toner Amarillo": "", "Toner Magenta": "", "Toner Cian": "", "Toner Negro": "", "Kit Alim.": "", 'Estado': '', 'Marca de Tiempo': ""}

        print(f"Procesando URL: {url}")
        driver = configurar_driver(num_version)

        try:

            driver.get(url)

            # Esperar a que cargue el primer consumible
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "SupplyName0"))
            )

            toner_amarillo = driver.find_element(By.ID, "SupplyGauge0").text
            toner_magenta = driver.find_element(By.ID, "SupplyGauge1").text
            toner_cian = driver.find_element(By.ID, "SupplyGauge2").text
            toner_negro = driver.find_element(By.ID, "SupplyGauge3").text
            kit_alimentador = driver.find_element(By.ID, "SupplyGauge4").text

            '''
            print(f"TÃ³ner Amarillo: {toner_amarillo}")
            print(f"TÃ³ner Magenta: {toner_magenta}")
            print(f"TÃ³ner Cian: {toner_cian}")
            print(f"TÃ³ner Negro: {toner_negro}")
            print(f"Kit alimentador: {kit_alimentador}")
            '''

            return {
                "IP": ip,
                "Toner Amarillo": toner_amarillo,
                "Toner Magenta": toner_magenta,
                "Toner Cian": toner_cian,
                "Toner Negro": toner_negro,
                "Kit Alim.": kit_alimentador,
                'Estado': 'OK' if toner_amarillo else 'No disponible',
                'Marca de Tiempo': timestamp
            }
        except (NoSuchElementException):
            return {"IP": ip, "Toner Amarillo": "", "Toner Magenta": "", "Toner Cian": "", "Toner Negro": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except TimeoutException:
            return {"IP": ip, "Toner Amarillo": "", "Toner Magenta": "", "Toner Cian": "", "Toner Negro": "", "Kit Alim.": "", 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
        except WebDriverException:
            print(f"Timeout al intentar conectar con {url}")
            return {"IP": ip, "Toner Amarillo": "", "Toner Magenta": "", "Toner Cian": "", "Toner Negro": "", "Kit Alim.": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
        finally:
            driver.quit()

    # Leer las hojas del archivo Excel
    sheets = pd.read_excel(file_path, sheet_name=None)
    wb = load_workbook(file_path)
    column_widths = {sheet_name: get_column_widths(
        wb[sheet_name]) for sheet_name in wb.sheetnames}

    # Procesar la hoja 'Color Admin'
    df_original = sheets['Color Planta']
    df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(
            fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    # Verificar que las columnas 'IP' existan antes de hacer el merge
    if 'IP' not in df_original.columns or 'IP' not in df_results.columns:
        raise KeyError("'IP' column is missing in one of the DataFrames.")

    # Fusionar los resultados
    df_updated = df_original.merge(
        df_results, on='IP', how='left', suffixes=('', '_new')
    )

    # Restablecer columnas NaN
    columns = ['Toner Amarillo', 'Toner Magenta', 'Toner Cian',
               'Toner Negro', 'Kit Alim.', 'Estado', 'Marca de Tiempo']
    df_updated[columns] = df_updated[columns].fillna('')

    mask_ok = df_updated['Estado_new'] == 'OK'
    columns = ['Toner Amarillo', 'Toner Magenta',
               'Toner Cian', 'Toner Negro', 'Kit Alim.']

    # ðŸ”¹ Crear las columnas *_new si no existen
    for col in columns:
        new_col = f"{col}_new"
        if new_col not in df_updated.columns:
            df_updated[new_col] = None

    # ðŸ”¹ Actualizar valores
    for col in columns:
        df_updated[col] = df_updated[col].astype(str)
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok,
                                                      f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(
        df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(
        df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in columns +
                       ['Estado', 'Marca de Tiempo']]
    df_updated.drop(columns=[
                    col for col in columns_to_drop if col in df_updated.columns], inplace=True)

    # Guardar el DataFrame actualizado
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(
            writer, sheet_name='Color Planta', index=False)
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'Color Planta':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar fÃ³rmulas y formatos preservados
    formulas = preserve_formulas_and_formats(file_path)
    apply_formulas_and_formats(output_file, formulas)
    # NO FUNCIONA
    registrar_historico(file_path, df_updated)


def format_excel_sheets(file_path):
    wb = load_workbook(file_path)
    red_font = Font(color="FF0000")
    orange_font = Font(color="ff6f00")

    print("Aplicando formato a todas las hojas:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Procesando hoja: {sheet_name}")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell_value = str(cell.value).strip()
              # Comprobar si el valor es '0%' o '0' para aplicar el texto rojo
                if cell_value in ['0%', '0']:
                    cell.font = red_font
                    print(
                        f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}' en hoja {sheet_name}")

                # Comprobar si el valor es menor al 5% para aplicar el texto rojo
                elif cell_value.endswith('%') and float(cell_value[:-1]) < 5:
                    cell.font = red_font
                    print(
                        f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}' en hoja {sheet_name}")

                # Comprobar si el valor es menor al 5% para aplicar el texto naranja
                elif cell_value.endswith('%') and float(cell_value[:-1]) < 5:
                    cell.font = orange_font
                    print(
                        f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}' en hoja {sheet_name}")

        # Ajustar el ancho de las columnas para cada hoja
        column_widths = get_column_widths(ws)
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    # Asegurar que "HP Admin" estÃ© al principio
    if "HP Admin" in wb.sheetnames:
        wb.move_sheet("HP Admin", offset=-
                      wb.index(wb["HP Admin"]))

    wb.save(file_path)
    print("Formato aplicado y archivo guardado.")


def registrar_historico(output_file, df_actual):
    """
    Guarda los niveles actuales de tÃ³ner en una hoja llamada 'HistÃ³rico' 
    dentro del mismo archivo Excel, sin borrar los registros anteriores.
    """
    df_historico_nuevo = df_actual.copy()

    # ðŸš¨ ESTA ES LA MODIFICACIÃ“N CLAVE: AÃ‘ADIR LOS NUEVOS TÃ“NERS
    columnas_historico = ['Nombre', 'IP', 'Modelo', 'Toner Negro', 'Toner Cian',
                          'Toner Magenta', 'Toner Amarillo','Kit Mant.', 'Kit Alim.', 'Estado', 'Marca de Tiempo']

    # AsegÃºrate de seleccionar solo las columnas que realmente existen en el DataFrame actual
    columnas_a_seleccionar = [
        col for col in columnas_historico if col in df_historico_nuevo.columns]
    df_historico_nuevo = df_historico_nuevo[columnas_a_seleccionar]

    try:
        # Intenta leer el archivo y la hoja 'HistÃ³rico'
        try:
            # Abrir en modo 'a' (append) para modificar la hoja
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:

                # Intentar leer los datos existentes
                try:
                    existing_df = pd.read_excel(
                        output_file, sheet_name="HistÃ³rico")

                    # Asegurarse de que las columnas coincidan para la concatenaciÃ³n
                    # (Esto es Ãºtil si el historial antiguo no tenÃ­a las columnas de color)
                    for col in columnas_a_seleccionar:
                        if col not in existing_df.columns:
                            existing_df[col] = ''

                    # Combinar los datos antiguos con los nuevos
                    df_final = pd.concat(
                        [existing_df, df_historico_nuevo], ignore_index=True)

                except (ValueError, FileNotFoundError):
                    # La hoja 'HistÃ³rico' no existe, estÃ¡ vacÃ­a, o el archivo es nuevo.
                    df_final = df_historico_nuevo

                # Escribir la hoja combinada (sobrescribe solo la hoja HistÃ³rico)
                df_final.to_excel(writer, sheet_name="HistÃ³rico", index=False)

        except FileNotFoundError:
            # Si el archivo no existe, crearlo y agregar la hoja 'HistÃ³rico'
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_historico_nuevo.to_excel(
                    writer, sheet_name="HistÃ³rico", index=False)

        print(
            f"âœ… Registro histÃ³rico agregado ({len(df_historico_nuevo)} filas).")

    except Exception as e:
        print(f"âŒ OcurriÃ³ un error al registrar el histÃ³rico: {e}")


#input_file = r"C:\Users\jvargas\Downloads\Impresoras - final.xlsx"
input_file = r"G:\Unidades compartidas\InformÃ¡tica\Impresoras - final.xlsx"


procesar_impresoras_hp(input_file, input_file)
procesar_color_planta(input_file, input_file)
procesar_planta(input_file, input_file)
procesar_impresoras_hp_grandes(input_file, input_file)
procesar_color_admin(input_file, input_file)
format_excel_sheets(input_file)
