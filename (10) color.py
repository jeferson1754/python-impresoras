import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Obtener la fecha y hora actual
timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def format_ip(ip):
    if pd.isna(ip):
        return None
    
    ip = str(ip)  # Convertir a cadena
    if not ip.strip():
        return None

    ip = re.sub(r'\D', '', ip)
    
    if len(ip) == 12:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 11:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 10:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 9:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 8:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    elif len(ip) == 7:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    else:
        return ip


def clean_percentage(value):
    try:
        if isinstance(value, str):
            value = value.replace('%', '').strip()
        return f"{int(round(float(value)))}%"
    except ValueError:
        return ""

def fetch_data_from_url(ip, options):
    url = f"http://{ip}" if ip else None
    if not url:
        return {"IP": ip, "Toner Negro": "", "UI Negro": "" ,"Toner Cian": "", "UI Cian": "" ,"Toner Magenta": "", "UI Magenta": "" ,"Toner Amarillo": "", "UI Amarillo": "" ,'Estado': '', 'Marca de Tiempo': ""}
    
    print(f"Procesando URL: {url}")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.get(url)
        WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ruifw_MainFrm")))

        toner_negro = driver.find_element(By.XPATH, "//tr[@id='1']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2]").text
        ui_negro = driver.find_element(By.XPATH, "(//tr[@id='1']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2])[2]").text

        toner_cian = driver.find_element(By.XPATH, "//tr[@id='2']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2]").text
        ui_cian = driver.find_element(By.XPATH, "(//tr[@id='2']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2])[2]").text

        toner_magenta = driver.find_element(By.XPATH, "//tr[@id='3']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2]").text
        ui_magenta = driver.find_element(By.XPATH, "(//tr[@id='3']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2])[2]").text

        toner_amarillo = driver.find_element(By.XPATH, "//tr[@id='4']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2]").text
        ui_amarillo = driver.find_element(By.XPATH, "(//tr[@id='4']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2])[2]").text

        return {
            "IP": ip,
            "Toner Negro": clean_percentage(toner_negro),
            "UI Negro": clean_percentage(ui_negro),
            "Toner Cian": clean_percentage(toner_cian),
            "UI Cian": clean_percentage(ui_cian),
            "Toner Magenta": clean_percentage(toner_magenta),
            "UI Magenta": clean_percentage(ui_magenta),
            "Toner Amarillo": clean_percentage(toner_amarillo),
            "UI Amarillo": clean_percentage(ui_amarillo),
            'Estado': 'OK',
            'Marca de Tiempo': timestamp
        }
    except (NoSuchElementException, TimeoutException):
        return {"IP": ip, "Toner Negro": "", "UI Negro": "" ,"Toner Cian": "", "UI Cian": "","Toner Magenta": "", "UI Magenta": "","Toner Amarillo": "", "UI Amarillo": "" , 'Estado': 'No Disponible', 'Marca de Tiempo': timestamp}
    except WebDriverException:
        return {"IP": ip, "Toner Negro": "", "UI Negro": "" ,"Toner Cian": "", "UI Cian": "","Toner Magenta": "", "UI Magenta": "","Toner Amarillo": "", "UI Amarillo": "" , 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
    finally:
        driver.quit()

def get_column_widths(ws):
    widths = {}
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        widths[column] = max_length + 2  # Adding a bit more space
    return widths

def preserve_formulas_and_formats(input_file):
    wb = load_workbook(input_file, data_only=False)
    formulas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Verificar si la celda tiene fórmula
                    sheet_formulas[cell.coordinate] = cell.formula
                else:
                    sheet_formulas[cell.coordinate] = cell.value
        formulas[sheet_name] = sheet_formulas

    return formulas

def apply_formulas_and_formats(output_file, formulas):
    wb = load_workbook(output_file, data_only=False)

    for sheet_name, sheet_formulas in formulas.items():
        ws = wb[sheet_name]

        for cell_coord, formula in sheet_formulas.items():
            cell = ws[cell_coord]
            if formula is not None:
                if isinstance(formula, str) and formula.startswith('='):
                    cell.formula = formula
                else:
                    cell.value = formula

    wb.save(output_file)

if __name__ == "__main__":
    input_file = r'C:\Users\jvargas\Desktop\Impresoras - final.xlsx'
    output_file = input_file

    # Leer todas las hojas del archivo Excel
    sheets = pd.read_excel(input_file, sheet_name=None)
    
    # Guardar el ancho de las columnas para cada hoja
    column_widths = {}
    wb = load_workbook(input_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        column_widths[sheet_name] = get_column_widths(ws)

    # Obtener la primera hoja
    df_original = sheets['Impresoras a Color']
    
    # Formatear IPs y filtrar datos
    df_original['IP'] = df_original['IP'].astype(str).apply(lambda x: format_ip(x) if pd.notna(x) else x)
    df_filtered = df_original[df_original['IP'].notna()]

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(fetch_data_from_url, ip, options): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    df_results = pd.DataFrame(results)

    df_updated = df_original.merge(df_results, on='IP', how='left', suffixes=('', '_new'))

    mask_ok = df_updated['Estado_new'] == 'OK'

    for col in ['Toner Negro', 'UI Negro', 'Toner Cian', 'UI Cian', 'Toner Magenta', 'UI Magenta', 'Toner Amarillo', 'UI Amarillo']:
        df_updated.loc[mask_ok, col] = df_updated.loc[mask_ok, f'{col}_new'].apply(clean_percentage)

    df_updated['Estado'] = df_updated['Estado_new'].fillna(df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(df_updated['Marca de Tiempo'])

    columns_to_drop = [f'{col}_new' for col in ['Toner Negro', 'UI Negro', 'Toner Cian', 'UI Cian', 'Toner Magenta', 'UI Magenta', 'Toner Amarillo', 'UI Amarillo', 'Estado', 'Marca de Tiempo']]
    columns_to_drop = [col for col in columns_to_drop if col in df_updated.columns]
    df_updated = df_updated.drop(columns=columns_to_drop)

    # Guardar el DataFrame actualizado en la hoja correcta
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(writer, sheet_name='Impresoras a Color', index=False)

        # Escribir las demás hojas que ya estaban en el archivo
        for sheet_name, df_sheet in sheets.items():
            if sheet_name != 'Impresoras a Color':
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar formato rojo a los valores '0%' y '0'
    wb = load_workbook(output_file, data_only=False)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        red_font = Font(color="FF0000")

        print(f"Aplicando formato a la hoja: {sheet_name}")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell_value = str(cell.value).strip()
                if cell_value in ['0%', '0']:
                    cell.font = red_font
                    print(f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}'")

        # Ajustar el ancho de las columnas
        original_widths = column_widths.get(sheet_name, {})
        for col, width in original_widths.items():
            ws.column_dimensions[col].width = width

    wb.save(output_file)
    
    # Aplicar fórmulas y formatos preservados
    formulas = preserve_formulas_and_formats(input_file)
    apply_formulas_and_formats(output_file, formulas)
