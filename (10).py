import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Obtener la fecha y hora actual
timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def format_ip(ip):
    if pd.isna(ip) or not ip.strip():
        return None  # Retorna None si la IP está vacía o solo tiene espacios

    ip = re.sub(r'\D', '', ip)
    
    if len(ip) == 12:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 11:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 10:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 9:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 8:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    elif len(ip) == 7:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    else:
        return ip

def fetch_data_from_url(ip):
    url = f"http://{ip}" if ip else None
    if not url:
        return {"IP": ip, "Toner Negro": "", "UI Negro": "", 'Estado': '', 'Marca de Tiempo': ""}
    
    print(f"Procesando URL: {url}")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.get(url)
        WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ruifw_MainFrm")))

        # Intentar obtener el valor del contador negro y blanco
        try:
            black_and_white_counter = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table#toner_list td.tonervalue_number"))
            ).text
        except (NoSuchElementException, TimeoutException):
            black_and_white_counter = "0%"

        # Intentar obtener el valor del contador de color
        try:
            color_counter = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table#imagine_list td.tonervalue_number"))
            ).text
        except (NoSuchElementException, TimeoutException):
            color_counter = "0%"

        return {"IP": ip, "Toner Negro": black_and_white_counter, "UI Negro": color_counter, 'Estado': 'OK', 'Marca de Tiempo': timestamp}
    except WebDriverException:
        return {"IP": ip, "Toner Negro": "", "UI Negro": "", 'Estado': 'Fuera de Red', 'Marca de Tiempo': timestamp}
    finally:
        driver.quit()

def convert_to_percentage(value):
    if pd.isna(value) or value is None:
        return ""  # Deja la celda vacía
    elif isinstance(value, (int, float)):
        return f"{int(value * 100)}%"
    elif isinstance(value, str):
        try:
            # Convertir de cadena a número, manejando coma como punto decimal
            number = float(value.replace(',', '.'))
            return f"{int(number * 100)}%"
        except ValueError:
            return value
    else:
        return value

def get_column_widths(ws):
    widths = {}
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        widths[column] = max_length + 2  # Adding a bit more space
    return widths

def preserve_formulas_and_formats(input_file):
    wb = load_workbook(input_file, data_only=False)
    formulas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Verificar si la celda tiene fórmula
                    sheet_formulas[cell.coordinate] = cell.formula
                else:
                    sheet_formulas[cell.coordinate] = cell.value
        formulas[sheet_name] = sheet_formulas

    return formulas

def apply_formulas_and_formats(output_file, formulas):
    wb = load_workbook(output_file, data_only=False)

    for sheet_name, sheet_formulas in formulas.items():
        ws = wb[sheet_name]

        for cell_coord, formula in sheet_formulas.items():
            cell = ws[cell_coord]
            if formula is not None:
                if isinstance(formula, str) and formula.startswith('='):
                    cell.formula = formula
                else:
                    cell.value = formula

    wb.save(output_file)

if __name__ == "__main__":
    input_file = r'C:\Users\jvargas\Desktop\Impresoras - final.xlsx'
    output_file = input_file

    # Leer el archivo Excel original
    df_original = pd.read_excel(input_file, dtype={"Toner Negro": str, "UI Negro": str})
    
    # Formatear las IPs
    df_original['IP'] = df_original['IP'].astype(str).apply(lambda x: format_ip(x) if pd.notna(x) else x)
    df_filtered = df_original[df_original['IP'].notna()]

    # Configuración de Selenium
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    # Ejecutar la función en paralelo
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(fetch_data_from_url, ip): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]

    # Crear un DataFrame con los resultados obtenidos
    df_results = pd.DataFrame(results)

    # Combinar los resultados con el DataFrame original
    df_updated = df_original.merge(df_results, on='IP', how='left', suffixes=('', '_new'))

    # Asegurarse de que las columnas estén en el tipo correcto antes de asignar
    df_updated['Toner Negro'] = df_updated['Toner Negro'].fillna('').astype(str)
    df_updated['UI Negro'] = df_updated['UI Negro'].fillna('').astype(str)

    # Actualizar 'Toner Negro' y 'UI Negro' solo si el estado es 'OK'
    mask_ok = df_updated['Estado_new'] == 'OK'
    df_updated.loc[mask_ok, 'Toner Negro'] = df_updated.loc[mask_ok, 'Toner Negro_new']
    df_updated.loc[mask_ok, 'UI Negro'] = df_updated.loc[mask_ok, 'UI Negro_new']

    # Actualizar 'Estado' y 'Marca de Tiempo' para todos los casos
    df_updated['Estado'] = df_updated['Estado_new'].fillna(df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(df_updated['Marca de Tiempo'])

    # Eliminar columnas auxiliares
    columns_to_drop = ['Toner Negro_new', 'UI Negro_new', 'Estado_new', 'Marca de Tiempo_new']
    df_updated = df_updated.drop(columns=columns_to_drop)

    # Convertir valores decimales a porcentajes
    df_updated['Toner Negro'] = df_updated['Toner Negro'].apply(convert_to_percentage)
    df_updated['UI Negro'] = df_updated['UI Negro'].apply(convert_to_percentage)

    # Guardar el DataFrame actualizado en un archivo Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_updated.to_excel(writer, sheet_name='Sheet1', index=False)

    # Aplicar formato rojo a los valores '0%' y '0'
    wb = load_workbook(output_file, data_only=False)
    ws = wb.active

    # Define el formato de texto rojo
    red_font = Font(color="FF0000")

    # Imprimir información para depuración
    print("Aplicando formato a los valores '0%' y '0':")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell_value = str(cell.value).strip()
            if cell_value in ['0%', '0']:
                cell.font = red_font
                print(f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}'")

    # Ajustar el ancho de las columnas
    column_widths = get_column_widths(ws)
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Guardar el archivo Excel con los formatos aplicados
    wb.save(output_file)

    # Aplicar fórmulas y formatos preservados
    formulas = preserve_formulas_and_formats(input_file)
    apply_formulas_and_formats(output_file, formulas)
