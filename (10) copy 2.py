import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Obtener la fecha y hora actual
timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def format_ip(ip):
    if pd.isna(ip):
        return None
    
    ip = str(ip)  # Convertir a cadena
    if not ip.strip():
        return None

    ip = re.sub(r'\D', '', ip)
    
    if len(ip) == 12:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 11:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:9]}.{ip[9:]}"
    elif len(ip) == 10:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 9:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:8]}.{ip[8:]}"
    elif len(ip) == 8:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    elif len(ip) == 7:
        return f"{ip[:3]}.{ip[3:6]}.{ip[6:]}"
    else:
        return ip

def fetch_data_from_url(ip, is_color_printer=False):
    url = f"http://{ip}" if ip else None
    if not url:
        return {
            "IP": ip,
            "Toner Negro": "", 
            "UI Negro": "",
            'Estado': '', 
            'Marca de Tiempo': ""
        }
    
    print(f"Procesando URL: {url}")
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        driver.get(url)
        WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ruifw_MainFrm")))

        black_and_white_counter = "0%"
        color_counter = "0%"

        try:
            black_and_white_counter = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table#toner_list td.tonervalue_number"))
            ).text
        except (NoSuchElementException, TimeoutException):
            pass

        try:
            color_counter = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table#imagine_list td.tonervalue_number"))
            ).text
        except (NoSuchElementException, TimeoutException):
            pass

        toner_cian = ui_cian = ""
        if is_color_printer:
            try:
                toner_cian = driver.find_element(By.XPATH, "//tr[@id='2']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2]").text
                ui_cian = driver.find_element(By.XPATH, "(//tr[@id='2']/td[2]/table/tbody/tr/td/table/tbody/tr/td[2])[2]").text
            except (NoSuchElementException, TimeoutException):
                toner_cian = "0%"
                ui_cian = "0%"

        return {
            "IP": ip,
            "Toner Negro": black_and_white_counter,
            "UI Negro": color_counter,
            "Toner Cian": toner_cian,
            "UI Cian": ui_cian,
            'Estado': 'OK',
            'Marca de Tiempo': timestamp
        }
    except WebDriverException:
        return {
            "IP": ip,
            "Toner Negro": "",
            "UI Negro": "",
            'Estado': 'Fuera de Red',
            'Marca de Tiempo': timestamp
        }
    finally:
        driver.quit()

def process_printers(df_original, is_color_printer=False, sheet_name='Impresoras Normales'):
    df_filtered = df_original[df_original['IP'].notna()]
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {executor.submit(fetch_data_from_url, format_ip(ip), is_color_printer): ip for ip in df_filtered['IP']}
        results = [future.result() for future in as_completed(future_to_ip)]
    
    df_results = pd.DataFrame(results)

    # Asegurarse de que las columnas 'IP' en ambos DataFrames sean del mismo tipo
    df_original['IP'] = df_original['IP'].astype(str)
    df_results['IP'] = df_results['IP'].astype(str)

    # Seleccionar solo las columnas necesarias según el tipo de impresora
    if is_color_printer:
        df_results = df_results[['IP', 'Toner Negro', 'UI Negro', 'Toner Cian', 'UI Cian', 'Estado', 'Marca de Tiempo']]
    else:
        df_results = df_results[['IP', 'Toner Negro', 'UI Negro', 'Estado', 'Marca de Tiempo']]

    df_updated = df_original.merge(df_results, on='IP', how='left', suffixes=('', '_new'))
    
    # Procesar solo las columnas relevantes según el tipo de impresora
    df_updated['Toner Negro'] = df_updated['Toner Negro_new'].fillna('').astype(str).apply(convert_to_percentage)
    df_updated['UI Negro'] = df_updated['UI Negro_new'].fillna('').astype(str).apply(convert_to_percentage)

    if is_color_printer:
        # Solo agregar estas columnas si es una impresora a color
        df_updated['Toner Cian'] = df_updated['Toner Cian_new'].fillna('').astype(str).apply(convert_to_percentage)
        df_updated['UI Cian'] = df_updated['UI Cian_new'].fillna('').astype(str).apply(convert_to_percentage)
    
    df_updated['Estado'] = df_updated['Estado_new'].fillna(df_updated['Estado'])
    df_updated['Marca de Tiempo'] = df_updated['Marca de Tiempo_new'].fillna(df_updated['Marca de Tiempo'])

    # Eliminar columnas innecesarias
    columns_to_drop = ['Toner Negro_new', 'UI Negro_new', 'Estado_new', 'Marca de Tiempo_new']
    if is_color_printer:
        columns_to_drop += ['Toner Cian_new', 'UI Cian_new']

    df_updated = df_updated.drop(columns=columns_to_drop)

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_updated.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_formats():
    wb = load_workbook(output_file, data_only=False)
    for sheet_name in ['Impresoras Normales', 'Impresoras a Color']:
        ws = wb[sheet_name]
        red_font = Font(color="FF0000")

        print(f"Aplicando formato a los valores '0%' y '0' en la hoja {sheet_name}:")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell_value = str(cell.value).strip()
                if cell_value in ['0%', '0']:
                    cell.font = red_font
                    print(f"Formato aplicado a celda: {cell.coordinate}, Valor: '{cell_value}'")

        column_widths = get_column_widths(ws)
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    wb.save(output_file)

def convert_to_percentage(value):
    if pd.isna(value) or value is None:
        return ""  
    elif isinstance(value, (int, float)):
        return f"{int(value * 100)}%"
    elif isinstance(value, str):
        try:
            value = float(value.strip('%'))
            return f"{int(value)}%"
        except ValueError:
            return value
    return value

def get_column_widths(sheet):
    column_widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            column = cell.column_letter
            if column not in column_widths:
                column_widths[column] = len(str(cell.value)) if cell.value else 10
            else:
                column_widths[column] = max(column_widths[column], len(str(cell.value)) if cell.value else 10)
    return column_widths

if __name__ == "__main__":
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')

    input_file = r'C:\Users\jvargas\Desktop\Impresoras - final.xlsx'
    output_file = input_file
    
    wb = load_workbook(input_file, data_only=True)
    
    # Procesar impresoras normales
    df_normal = pd.read_excel(input_file, sheet_name='Impresoras Normales')
    process_printers(df_normal, is_color_printer=False, sheet_name='Impresoras Normales')

    # Procesar impresoras a color
    df_color = pd.read_excel(input_file, sheet_name='Impresoras a Color')
    process_printers(df_color, is_color_printer=True, sheet_name='Impresoras a Color')

    # Aplicar formatos
    apply_formats()