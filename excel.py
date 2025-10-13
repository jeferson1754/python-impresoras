import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from contextlib import contextmanager

# Get current timestamp
timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# Determine the correct ChromeDriver version
# Use a more stable, general approach to avoid hardcoding versions
try:
    CHROMEDRIVER_PATH = ChromeDriverManager().install()
except ValueError as e:
    # Fallback to a hardcoded version if the automatic detection fails
    print(
        f"⚠️ Error with automatic ChromeDriver install: {e}. Falling back to a hardcoded version.")
    CHROMEDRIVER_PATH = ChromeDriverManager(
        driver_version="140.0.7339.208").install()  # Example version

# --- Driver Configuration ---


@contextmanager
def get_driver():
    """Context manager for setting up and tearing down the Selenium WebDriver."""
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    # options.add_argument('--window-size=1920,1080') # Recommended for headless
    driver = None
    try:
        # This line should automatically find and install the correct driver for your browser
        CHROMEDRIVER_PATH = ChromeDriverManager().install()
        print(f"✅ ChromeDriver installed/updated to version compatible with your Chrome.")
    except Exception as e:
        # If automatic installation fails, print an error and you may need to manually update ChromeDriverManager
        print(
            f"🚫 FATAL ERROR: Could not automatically install ChromeDriver. Check your internet connection or update the 'webdriver-manager' package. Error: {e}")
        # Re-raise the exception to stop the script if the driver can't be set up.
        raise
    finally:
        if driver:
            driver.quit()

# --- Helper Functions ---


def format_ip(ip):
    """Formats a raw string into a standard IP address format."""
    if pd.isna(ip) or not str(ip).strip():
        return None

    ip_str = re.sub(r'\D', '', str(ip))

    if len(ip_str) in [12, 11, 10, 9, 8, 7]:
        # This formatting logic is highly specific and might need a more general rule.
        # However, keeping it as is to match the original code's intent.
        parts = []
        if len(ip_str) >= 3:
            parts.append(ip_str[0:3])
        if len(ip_str) >= 6:
            parts.append(ip_str[3:6])
        if len(ip_str) >= 8:
            parts.append(ip_str[6:8])
        if len(ip_str) >= 9:
            parts.append(ip_str[8:9])
        if len(ip_str) >= 10:
            parts.append(ip_str[8:10])
        if len(ip_str) >= 12:
            parts.append(ip_str[9:12])

        # A more robust and generalized IP formatter should be considered.
        return ".".join(parts)
    else:
        return str(ip)


def clean_percentage(value: str) -> str:
    """Cleans and formats a string value to a percentage string."""
    try:
        if isinstance(value, str):
            value = value.replace('%', '').strip()
        if value:
            return f"{int(round(float(value)))}%"
    except (ValueError, TypeError):
        pass  # Return empty string implicitly
    return ""


def get_column_widths(ws):
    """Calculates optimal column widths for an openpyxl worksheet."""
    widths = {}
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (ValueError, TypeError):
                pass
        widths[column] = max_length + 2
    return widths


def format_excel_sheets(file_path):
    """Applies formatting and auto-fits columns for all sheets in an Excel file."""
    try:
        wb = load_workbook(file_path)
        red_font = Font(color="FF0000")
        orange_font = Font(color="ff6f00")

        print("Aplicando formato a todas las hojas...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Procesando hoja: {sheet_name}")

            # Apply font formatting
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell_value = str(cell.value).strip()
                    if cell_value.endswith('%'):
                        try:
                            percentage = float(cell_value[:-1])
                            if percentage < 5:
                                cell.font = red_font
                            elif percentage < 10:
                                cell.font = orange_font
                        except (ValueError, TypeError):
                            pass

            # Adjust column widths
            column_widths = get_column_widths(ws)
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

        # Move 'HP Admin' to the first position
        if "HP Admin" in wb.sheetnames:
            wb.move_sheet(wb["HP Admin"], offset=-wb.index(wb["HP Admin"]))

        wb.save(file_path)
        print("✅ Formato aplicado y archivo guardado.")
    except Exception as e:
        print(f"🚫 Error al aplicar formato a las hojas: {e}")


def registrar_historico(output_file, df_actual, sheet_name):
    """
    Adds a historical record of printer data to a 'Histórico' sheet.

    Args:
        output_file (str): Path to the Excel file.
        df_actual (pd.DataFrame): The DataFrame with the current data.
        sheet_name (str): The name of the sheet being processed.
    """
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df_historico = df_actual.copy()
    df_historico["Fecha de registro"] = fecha_actual
    df_historico["Hoja de origen"] = sheet_name

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # Check if the 'Histórico' sheet exists
            if 'Histórico' in writer.book.sheetnames:
                # Read existing data and append
                existing_df = pd.read_excel(
                    output_file, sheet_name="Histórico")
                df_final = pd.concat(
                    [existing_df, df_historico], ignore_index=True)

                # Overwrite the sheet with the combined data
                df_final.to_excel(writer, sheet_name="Histórico", index=False)
            else:
                # Create the sheet for the first time
                df_historico.to_excel(
                    writer, sheet_name="Histórico", index=False)

        print(f"✅ Registro histórico para '{sheet_name}' agregado.")

    except Exception as e:
        print(f"⚠️ Error al registrar histórico para '{sheet_name}': {e}")


# --- Main Processing Logic (Refactored) ---
def fetch_data_from_url(ip, consumables_to_check):
    """Fetches printer data from a given IP address."""
    url = f"http://{ip}" if ip else None
    if not url:
        return {"IP": ip, "Estado": "No Disponible"}

    print(f"Procesando URL: {url}")

    results = {"IP": ip, "Estado": "OK", "Marca de Tiempo": timestamp}

    with get_driver() as driver:
        try:
            driver.get(url)
            wait = WebDriverWait(driver, 10)

            # Wait for the first consumable element to be present
            wait.until(EC.presence_of_element_located((By.ID, "SupplyName0")))

            # Extract data for each specified consumable
            for i in range(len(consumables_to_check)):
                key = consumables_to_check[i]
                try:
                    results[key] = driver.find_element(
                        By.ID, f"SupplyGauge{i}").text
                except NoSuchElementException:
                    results[key] = ""  # Consumable not found

        except (NoSuchElementException, TimeoutException):
            results["Estado"] = "No Disponible"
        except WebDriverException:
            results["Estado"] = "Fuera de Red"
        except Exception as e:
            results["Estado"] = f"Error: {e}"
            print(f"🚫 Error inesperado para {url}: {e}")

    return results


def process_printers(file_path, output_file, sheet_name, columns_to_check, num_workers=5):
    """
    Generalized function to process different printer types based on sheet name.
    """
    try:
        print(f"🚀 Iniciando procesamiento para la hoja '{sheet_name}'...")

        # Read the sheet and format IPs
        df_original = pd.read_excel(file_path, sheet_name=sheet_name)
        df_original['IP'] = df_original['IP'].astype(str).apply(format_ip)
        df_filtered = df_original[df_original['IP'].notna()]

        results = []
        # Use a ThreadPoolExecutor for concurrent web scraping
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            future_to_ip = {
                executor.submit(fetch_data_from_url, ip, columns_to_check): ip
                for ip in df_filtered['IP']
            }

            # 💡 FIX: Iterate over results and handle exceptions for each future
            for future in as_completed(future_to_ip):
                ip = future_to_ip[future]
                try:
                    results.append(future.result())
                except Exception as exc:
                    print(f'🚫 IP {ip} generó una excepción: {exc}')
                    # Append a default failure result to maintain structure
                    default_fail = {"IP": ip, "Estado": "Error Fatal"}
                    for col in columns_to_check:
                        default_fail[col] = ""
                    results.append(default_fail)

        df_results = pd.DataFrame(results)

        # ... (rest of the processing logic remains the same) ...
        # Merge results with original DataFrame
        df_updated = df_original.merge(
            df_results, on='IP', how='left', suffixes=('', '_new')
        )

        # Update the original columns with the new data
        for col in columns_to_check:
            df_updated[f'{col}_new'] = df_updated[f'{col}_new'].apply(
                clean_percentage)
            df_updated[col].update(df_updated[f'{col}_new'])

        # Update Status and Timestamp columns
        df_updated['Estado'].update(df_updated['Estado_new'])
        df_updated['Marca de Tiempo'].update(df_updated['Marca de Tiempo_new'])

        # Clean up temporary columns
        columns_to_drop = [
            f'{col}_new' for col in columns_to_check] + ['Estado_new', 'Marca de Tiempo_new']
        df_updated.drop(
            columns=[c for c in columns_to_drop if c in df_updated.columns], inplace=True)

        # Save the updated DataFrame to the Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_updated.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"✅ Procesamiento para '{sheet_name}' completado.")

        # Register historical data
        registrar_historico(output_file, df_updated, sheet_name)

    except FileNotFoundError:
        print(f"🚫 Error: El archivo '{file_path}' no se encontró.")
    except Exception as e:
        # Catch any remaining exception outside the futures loop
        print(f"🚫 Error general al procesar la hoja '{sheet_name}': {e}")


def preserve_formulas_and_formats(input_file):
    wb = load_workbook(input_file, data_only=False)
    formulas = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = {cell.coordinate: cell.formula if cell.data_type ==
                          'f' else cell.value for row in ws.iter_rows() for cell in row}
        formulas[sheet_name] = sheet_formulas
    return formulas


def apply_formulas_and_formats(output_file, formulas):
    wb = load_workbook(output_file, data_only=False)
    for sheet_name, sheet_formulas in formulas.items():
        ws = wb[sheet_name]
        for cell_coord, formula in sheet_formulas.items():
            cell = ws[cell_coord]
            if formula is not None:
                if isinstance(formula, str) and formula.startswith('='):
                    cell.formula = formula
                else:
                    cell.value = formula
    wb.save(output_file)


def main():
    """Main function to run the entire script."""
    input_file = r"C:\Users\jvargas\Downloads\Impresoras - final.xlsx"

    # Define the sheets and the consumables to check for each
    sheets_config = {
        'HP Admin': ['Toner Negro', 'Kit Mant.', 'Kit Alim.'],
        'HP Planta': ['Toner Negro'],
        'HP Planta - 2': ['Toner Negro', 'Kit Alim.'],
        'Color Admin': ['Toner Negro', 'Toner Cian', 'Toner Magenta', 'Toner Amarillo'],
        'Color Planta': ['Toner Amarillo', 'Toner Magenta', 'Toner Cian', 'Toner Negro', 'Kit Alim.']
    }

    # Preserve formulas and formats before processing
    formulas_and_formats = {}
    try:
        formulas_and_formats = preserve_formulas_and_formats(input_file)
    except FileNotFoundError:
        print("⚠️ El archivo de entrada no se encontró. Creando un nuevo archivo.")

    # Process each sheet
    for sheet_name, columns in sheets_config.items():
        process_printers(input_file, input_file, sheet_name, columns)

    # Apply preserved formulas and formatting
    if formulas_and_formats:
        apply_formulas_and_formats(input_file, formulas_and_formats)

    # Format the Excel sheets (font color, column width)
    format_excel_sheets(input_file)


if __name__ == "__main__":
    main()
